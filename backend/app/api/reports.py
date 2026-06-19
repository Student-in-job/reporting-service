from io import BytesIO
from uuid import UUID

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.database import get_db
from app.api.deps import get_current_user
from app.models.user import User
from app.models.report import Report
from app.models.datasource import Datasource
from app.schemas.report import ReportListOut, ReportListItem, ReportDataRequest, ReportDataOut
from app.services.report_executor import report_executor

router = APIRouter()


@router.get("", response_model=ReportListOut)
async def list_reports(
    group: str | None = None,
    db: AsyncSession = Depends(get_db),
    _: User = Depends(get_current_user),
):
    query = select(Report).where(Report.is_active.is_(True)).order_by(Report.group, Report.name)
    if group:
        query = query.where(Report.group == group)

    result = await db.execute(query)
    items = []
    for r in result.scalars().all():
        config = r.config or {}
        items.append(
            ReportListItem(
                id=r.id,
                slug=r.slug,
                name=r.name,
                description=r.description,
                group=r.group,
                type=r.type,
                filters=config.get("filters", []),
            )
        )
    return ReportListOut(reports=items)


async def _get_report_and_datasource(report_id: str, db: AsyncSession):
    try:
        uid = UUID(report_id)
        query = select(Report).where(Report.id == uid, Report.is_active.is_(True))
    except ValueError:
        query = select(Report).where(Report.slug == report_id, Report.is_active.is_(True))

    result = await db.execute(query)
    report = result.scalar_one_or_none()
    if report is None:
        raise HTTPException(status_code=404, detail="Report not found")

    ds_result = await db.execute(select(Datasource).where(Datasource.id == report.datasource_id))
    datasource = ds_result.scalar_one_or_none()
    if datasource is None:
        raise HTTPException(status_code=500, detail="Datasource not configured")

    return report, datasource


@router.post("/{report_id}/data", response_model=ReportDataOut)
async def execute_report(
    report_id: str,
    body: ReportDataRequest,
    db: AsyncSession = Depends(get_db),
    _: User = Depends(get_current_user),
):
    report, datasource = await _get_report_and_datasource(report_id, db)
    params = {"date_from": body.date_from, "date_to": body.date_to, "extra_filters": body.filters}

    try:
        return await report_executor.execute(report, datasource, params)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=504, detail=f"Query execution failed: {e}")


@router.post("/{report_id}/export")
async def export_report(
    report_id: str,
    body: ReportDataRequest,
    db: AsyncSession = Depends(get_db),
    _: User = Depends(get_current_user),
):
    report, datasource = await _get_report_and_datasource(report_id, db)
    params = {"date_from": body.date_from, "date_to": body.date_to, "extra_filters": body.filters}

    try:
        result = await report_executor.execute(report, datasource, params)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=504, detail=f"Query execution failed: {e}")

    wb = Workbook()
    ws = wb.active
    ws.title = report.name[:31]

    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    # Header row
    for col_idx, col_def in enumerate(result.columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_def.label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Data rows
    for row_idx, row_data in enumerate(result.data, 2):
        for col_idx, col_def in enumerate(result.columns, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=row_data.get(col_def.key))
            cell.border = thin_border

    # Auto-width
    for col_idx, col_def in enumerate(result.columns, 1):
        max_len = len(col_def.label)
        for row_idx in range(2, len(result.data) + 2):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val is not None:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 3, 50)

    # Freeze header
    ws.freeze_panes = "A2"

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"{report.slug}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
